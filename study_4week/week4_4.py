import openpyxl

wb=openpyxl.Workbook()

sheet1=wb.active

sheet1.title = "1st sheet"

sheet2 = wb.create_sheet("2nd sheet")

for i in range(1,10):
    sheet1.cell(row=i, column=1).value = i
    sheet2.cell(row=1, column=i).value = i

wb.save('test.xlsx')

